import streamlit as st
import pandas as pd
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
import os
import time
import requests
from io import BytesIO
from openpyxl.styles import Border, Side,Alignment


def convert_pdf_to_excel(pdf_file):

    rows = []
    firstcolumn=''
    # Function to extract text from a PDF and process it into a tabular format
    def extract_data(pdf_file,pdf_name_withoutextension):
        # Open the PDF file
        with open(pdf_file, 'rb') as file:
            reader = PyPDF2.PdfReader(file)

            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                text = page.extract_text()



                # Split the text into lines and process each line
                for line in text.split('\n'):
            
                    rows.append(line)
        
                
            sublists = []

            i = 3  # Index to track position in the list

            while i < len(rows):
                # Check if the current value is "Department Name"
                if rows[i] == 'Department Name':
                    # Skip next 6 values (including the current one)
                    
                    department=rows[i+1]
                    department=department.strip()
                    section=rows[i+5]
                    sem=rows[i+2]

                    i += 7
                    print(department)
                    print(sem)
                    print(section)
                else:
                    # Extract a sublist of next 3 elements
                    sublist = rows[i:i + 3]
                    sublists.append(sublist)
                    i += 3  # Move the index to the next set of 3 elements
            
            df = pd.DataFrame(sublists)
            firstcolumn='Student Name '+department+' '+sem+'  '+section
            firstcolumn=firstcolumn.upper()
            df.to_excel(pdf_name_withoutextension,header=[firstcolumn,'USN','PASSWORD'], index=False)
            return rows,firstcolumn

    # Path to your PDF file

    pdf_file_name= pdf_file
    pdf_name_withoutextension=pdf_file_name.split('.')[0]+'.xlsx'


    # Extract table data
    table_data,firstcolumn = extract_data(pdf_file_name,pdf_name_withoutextension)


    file_path =pdf_name_withoutextension  # Replace with your file path
    df = pd.read_excel(file_path, engine='openpyxl')

    # Add a Serial Number column
    df.insert(0, 'Sl. No.', range(1, 1 + len(df)))

    # Save the modified DataFrame back to Excel
    df.to_excel(file_path, index=False)


    col1 = firstcolumn  # Replace with the name of the first column
    col2 = 'USN'  # Replace with the name of the second column

    # Swap the columns
    df[col1], df[col2] = df[col2].copy(), df[col1].copy()
    df.to_excel(file_path, index=False)


    file_path = pdf_name_withoutextension # Replace with your file path
    workbook = load_workbook(file_path)

    # Select the active worksheet
    worksheet = workbook.active
    align_center = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D']:
        for row in range(1, worksheet.max_row + 1):
            cell = worksheet[f'{col}{row}']
            cell.alignment = align_center
    thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Apply the border to each cell
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
        # Set the height of each row
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 25
        workbook.save(pdf_name_withoutextension)



    file_path = pdf_name_withoutextension  # Replace with your file path

    def get_file_content_as_bytes(path):
        with open(path, "rb") as file:
            return file.read()

   
    file_content = get_file_content_as_bytes(file_path)
    st.download_button(
        label='Download Excel File',
        data=file_content,
        file_name=file_path,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


st.title('PDF to Excel Converter')

uploaded_files = st.file_uploader("Choose files", accept_multiple_files=True)
for uploaded_file in uploaded_files:

    if uploaded_file is not None:
        file_name = uploaded_file.name

        # Define the file path (current directory in this case)
        file_path = os.path.join(os.getcwd(), file_name)

        # Write the file to the current directory
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        st.success(f'File "{file_name}" saved at "{file_path}"')

    excel_file = convert_pdf_to_excel(file_name)


    









